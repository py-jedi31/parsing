from oauth2client.service_account import ServiceAccountCredentials
import gspread
import vk_api
import json
from itertools import zip_longest
import xlrd
from openpyxl import load_workbook, Workbook
import os
import pandas as pd
import colorama
from colorama import Fore

colorama.init()

def get_params(file_path):
    data = json.load(open(file_path, encoding='utf8'))
    params = data['settings'].values()
    return params

class GoogleSheetParser:

    def __init__(self, file_path, token, login, password):
        self.file_path = file_path
        self.token = token
        self.login = login
        self.password = password

    def read_google_sheet(self, sheet_names):
        # подключаемые API сервисы Google
        scope = ('https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive')
        # учётные данные из .json файла от Google
        credentials = ServiceAccountCredentials.from_json_keyfile_name(self.file_path)
        # авторизация в Google Sheets
        auth = gspread.authorize(credentials)
        # несколько листов надо!
        wb_lists = []
        for wb in sheet_names:
            wb_list = auth.open(wb).worksheets()
            wb_lists.append(wb_list)

        return wb_lists

    def send_in_vk(self, user_id):
        # создание сессии ВКонтакте
        vk_session = vk_api.VkApi(login=self.login, password=self.password, token=self.token)
        vk = vk_session.get_api()
        message_text = 'Ваше задание выполнено.'
        vk.messages.send(user_id=user_id, message=message_text, random_id=0)

    def check_google_sheet(self, wb_lists):
        if not os.path.exists('diary0.xlsx'):
            for sheet_index, wb_list in enumerate(wb_lists):
                dfs = []
                # сохранение листов в список
                for wb in wb_list:
                    df = pd.DataFrame(wb.get_all_records())
                    dfs.append(df)
                # сохранение листов в один документ
                _wb = Workbook()
                _wb.save(f'diary{sheet_index}.xlsx')
                with pd.ExcelWriter(f'diary{sheet_index}.xlsx', mode='a') as writer:
                    for index, _df in enumerate(dfs):
                        _df.to_excel(writer, sheet_name=f'Sheet{index}')
        else:
            for sheet_index, wb_list in enumerate(wb_lists):
                dfs = []
                for wb in wb_list:
                    df = pd.DataFrame(wb.get_all_records())
                    dfs.append(df)

                _wb = Workbook()
                _wb.save(f'new_diary{sheet_index}.xlsx')

                with pd.ExcelWriter(f'new_diary{sheet_index}.xlsx', mode='a') as writer:
                    for index, _df in enumerate(dfs):
                        _df.to_excel(writer, sheet_name=f'Sheet{index}')

                # процесс сравнения
                rb1 = xlrd.open_workbook(f'diary{sheet_index}.xlsx')
                rb2 = xlrd.open_workbook(f'new_diary{sheet_index}.xlsx')
                all_worksheets = max(len(rb1.sheet_names()), len(rb2.sheet_names()))

                for index in range(1, all_worksheets):

                    sheet1 = rb1.sheet_by_index(index)
                    sheet2 = rb2.sheet_by_index(index)

                    for rownum in range(max(sheet1.nrows, sheet2.nrows)):

                        if rownum < sheet1.nrows:
                            row_rb1 = sheet1.row_values(rownum)
                            row_rb2 = sheet2.row_values(rownum)

                            for colnum, (c1, c2) in enumerate(zip_longest(row_rb1, row_rb2)):
                                n = 0
                                if type(c1) != type(c2):
                                    #print("Row {} Col {} - {} != {}".format(rownum + 1, colnum + 1, c1, c2))
                                    n += 1
                                    for i in range(1, n + 1):
                                        self.send_in_vk(user_id=str(int(sheet1.cell_value(rownum, sheet1.ncols - 1))))
                                        print(Fore.GREEN + f'Отправлено {str(int(sheet1.cell_value(rownum, sheet1.ncols - 1)))}')
                        else:
                            print(Fore.CYAN + 'Изменения отсутствуют.')

                wb = load_workbook(f'new_diary{sheet_index}.xlsx')
                wb.save(f'diary{sheet_index}.xlsx')
                os.remove(f'new_diary{sheet_index}.xlsx')



# запуск скрипта
if __name__ == '__main__':
    try:
        params = get_params('settings.json')
        GOOGLE_JSON_FILE, VK_TOKEN, LOGIN, PASSWORD, NAME_SHEETS = \
            (param for param in params)
        bot = GoogleSheetParser(GOOGLE_JSON_FILE, VK_TOKEN, LOGIN, PASSWORD)
        wks = bot.read_google_sheet(NAME_SHEETS)
        bot.check_google_sheet(wks)

    except Exception as error:
        print(Fore.RED + str(error))
    finally:
        print(Fore.CYAN + 'Программа завершена. Нажмите enter для выхода.')
        io = input()
